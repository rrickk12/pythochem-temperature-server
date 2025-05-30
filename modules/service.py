# modules/service.py

import logging
from db_ops.db_manager import DatabaseManager
from modules.reader import DataReader
from modules.report import ReportGenerator
import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

class SensorService:
    """
    Service layer to interact with sensor data.
    Abstracts underlying database operations and report generation.
    """
    def __init__(self, db_url='sqlite:///ble_data.db'):
        self.db_manager = DatabaseManager(db_url)
        self.data_reader = DataReader(self.db_manager)
        self.report_generator = ReportGenerator(self.data_reader)

    def get_all_sensors(self):
        sensors = self.db_manager.get_all_sensors()
        logger.debug("Service: Retrieved sensors: %s", sensors)
        return sensors

    def get_sensor_report(self, mac, start_timestamp, end_timestamp):
        report = self.report_generator.generate_sensor_report(mac, start_timestamp, end_timestamp)
        logger.debug("Service: Generated report for sensor %s", mac)
        return report

    def update_sensor_alert_policy(self, mac, temp_min=None, temp_max=None, humidity_min=None, humidity_max=None):
        self.db_manager.set_alert_policy(mac, temp_min, temp_max, humidity_min, humidity_max)
        logger.debug("Service: Updated alert policy for sensor %s", mac)
        return {"status": "Alert policy updated", "mac": mac}

    def update_sensor_schedule_policy(self, mac, delta_time):
        self.db_manager.set_schedule_policy(mac, delta_time)
        logger.debug("Service: Updated schedule policy for sensor %s", mac)
        return {"status": "Schedule policy updated", "mac": mac}

    # Use este método só se quiser rota conjunta!
    def update_sensor_policies(self, mac, temp_min=None, temp_max=None, humidity_min=None, humidity_max=None, delta_time=None):
        if any([temp_min, temp_max, humidity_min, humidity_max]):
            self.db_manager.set_alert_policy(mac, temp_min, temp_max, humidity_min, humidity_max)
            logger.debug("Service: Updated alert policy for sensor %s", mac)
        if delta_time is not None:
            self.db_manager.set_schedule_policy(mac, delta_time)
            logger.debug("Service: Updated schedule policy for sensor %s", mac)
        return {"status": "Policies updated", "mac": mac}

    def rename_sensor(self, mac, name):
        result = self.db_manager.rename_sensor(mac, name)
        logger.debug("Service: Renamed sensor %s to '%s'", mac, name)
        return {"status": "Sensor renamed" if result else "Sensor not found", "mac": mac, "name": name}

# def export_sensor_data(self, mac, fr, to, interval=None):
#     # Usa direto a leitura agregada do banco (ReadScheduled)
#     reads = self.db_manager.get_scheduled_reads(mac, fr, to)
#     return [
#         {
#             "timestamp": r.timestamp,
#             "avg_temp": r.avg_temp,
#             "avg_hum": r.avg_hum,
#             "min_temp": r.min_temp,
#             "max_temp": r.max_temp,
#             "min_hum": r.min_hum,
#             "max_hum": r.max_hum,
#         }
#         for r in reads
#     ]
    def export_sensor_data(self, mac, fr, to, interval=None):
        from datetime import datetime, timedelta

        fr_dt = datetime.fromisoformat(fr)
        to_dt = datetime.fromisoformat(to)
        interval_h = int(interval) if interval else 1

        # Pega todos os CLEAN daquele sensor dentro do período
        reads = self.db_manager.get_latest_clean_reads(mac, limit=10000)
        reads = [
            r for r in reads
            if fr_dt <= datetime.fromisoformat(r.timestamp) <= to_dt
        ]
        if not reads:
            return []

        # Ordena por timestamp
        reads.sort(key=lambda r: r.timestamp)

        # Agrupa por intervalos de N horas
        result = []
        group = []
        group_start = datetime.fromisoformat(reads[0].timestamp)
        for r in reads:
            ts = datetime.fromisoformat(r.timestamp)
            # Se passou do intervalo, exporta o grupo e começa novo
            if (ts - group_start).total_seconds() >= interval_h * 3600:
                if group:
                    # Agrega o grupo
                    result.append(_aggregate_group(group, group_start))
                group = []
                group_start = ts
            group.append(r)
        # Último grupo
        if group:
            result.append(_aggregate_group(group, group_start))
        return result
    
    def export_all_sensors_data(self, fr, to, interval=None):
        """
        Exporta os dados agregados de todos os sensores no período e agrupamento informados.
        Retorna um dicionário {mac: [leituras...], ...}
        """
        sensors = self.db_manager.get_all_sensors()
        result = {}
        for sensor in sensors:
            mac = sensor.mac
            data = self.export_sensor_data(mac, fr, to, interval)
            if data:  # Só inclui se houver dados
                result[mac] = data
        return result

    def export_all_to_excel(json_data):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for mac, rows in json_data.items():
            ws = wb.create_sheet(title=mac[:31])  # Sheet title max 31 chars
            # Cabeçalho
            headers = ["timestamp", "avg_temp", "avg_hum", "min_temp", "max_temp", "min_hum", "max_hum"]
            ws.append(headers)
            for r in rows:
                ws.append([
                    r.get("timestamp"),
                    r.get("avg_temp"),
                    r.get("avg_hum"),
                    r.get("min_temp"),
                    r.get("max_temp"),
                    r.get("min_hum"),
                    r.get("max_hum")
                ])
            # Auto width
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_length, 10)

        # Salvar em memória
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    
    def export_one_to_excel(sensor_name, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = sensor_name[:31]
        headers = ["timestamp", "avg_temp", "avg_hum", "min_temp", "max_temp", "min_hum", "max_hum"]
        ws.append(headers)
        for r in rows:
            ws.append([
                r.get("timestamp"),
                r.get("avg_temp"),
                r.get("avg_hum"),
                r.get("min_temp"),
                r.get("max_temp"),
                r.get("min_hum"),
                r.get("max_hum")
            ])
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_length, 10)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

def _aggregate_group(group, group_start):
    """
    Faz agregação dos dados do grupo:
      - Média dos valores
      - Min/max
      - Última leitura válida
      - Horário da última leitura
    """
    # Ordene o grupo pelo timestamp (garante que o último é realmente o último)
    group_sorted = sorted(group, key=lambda g: g.timestamp)
    last = group_sorted[-1]
    first = group_sorted[0]

    return {
        "timestamp": group_start.isoformat(timespec='minutes'),
        "avg_temp": sum([g.avg_temp for g in group if g.avg_temp is not None]) / len(group),
        "avg_hum": sum([g.avg_hum for g in group if g.avg_hum is not None]) / len(group),
        "min_temp": min([g.min_temp for g in group if g.min_temp is not None]),
        "max_temp": max([g.max_temp for g in group if g.max_temp is not None]),
        "min_hum": min([g.min_hum for g in group if g.min_hum is not None]),
        "max_hum": max([g.max_hum for g in group if g.max_hum is not None]),
        "last_temp": last.avg_temp,
        "last_hum": last.avg_hum,
        "last_timestamp": last.timestamp,
        "first_temp": first.avg_temp,
        "first_hum": first.avg_hum,
        "first_timestamp": first.timestamp,
    }

# Singleton para import fácil
sensor_service = SensorService()
