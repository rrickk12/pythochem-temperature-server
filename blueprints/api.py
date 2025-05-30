import logging
from flask import Blueprint, request, jsonify, abort
from modules.service import sensor_service
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def get_limits_for_sensor(sensor_mac, db_manager):
    policy = db_manager.get_alert_policy(sensor_mac)
    if not policy:
        return {}
    return {
        "temp_min": policy.temp_min,
        "temp_max": policy.temp_max,
        "hum_min": policy.humidity_min,
        "hum_max": policy.humidity_max,
    }
def highlight_cell(cell, value, lim_min, lim_max):
    # Sem limite configurado
    if lim_min is None and lim_max is None:
        return
    yellow = PatternFill(start_color="FFF475", end_color="FFF475", fill_type="solid")
    red    = PatternFill(start_color="FF8A80", end_color="FF8A80", fill_type="solid")
    try:
        value = float(value)
    except (TypeError, ValueError):
        return
    # Fora do limite
    if lim_min is not None and value < lim_min:
        cell.fill = red
    elif lim_max is not None and value > lim_max:
        cell.fill = red
    elif (lim_min is not None and value == lim_min) or (lim_max is not None and value == lim_max):
        cell.fill = yellow

def write_sensor_sheet(ws, rows, headers, limits=None):
    ws.append([h[1] for h in headers])

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Linhas zebradas
    fill1 = PatternFill(start_color="F8FAFF", end_color="F8FAFF", fill_type="solid")
    fill2 = PatternFill(start_color="E5EDF7", end_color="E5EDF7", fill_type="solid")

    # Mapeamento de colunas
    col_idx = {k: i+1 for i, (k, _) in enumerate(headers)}

    for i, row in enumerate(rows, start=2):
        # Quebra timestamp em data/hora (se os headers começarem com "date" e "time")
        if "timestamp" in row:
            ts = row["timestamp"]
            if "T" in ts:
                date, time = ts.split("T")
                time = time[:5]
            else:
                date, time = ts, ""
        else:
            date, time = "", ""

        # Prepara valores conforme headers (primeiros dois são date/time, resto igual)
        values = []
        for j, (key, _) in enumerate(headers):
            if key == "date":
                values.append(date)
            elif key == "time":
                values.append(time)
            elif key == "last_timestamp":
                last_ts = row.get("last_timestamp", "")
                values.append(last_ts.split("T")[1][:5] if "T" in last_ts else last_ts)
            else:
                values.append(row.get(key, ""))

        ws.append(values)
        fill = fill1 if i % 2 == 0 else fill2

        for key, idx in col_idx.items():
            cell = ws.cell(row=i, column=idx)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            # Formatação numérica nas colunas (exceto Data/Hora)
            if key not in ("date", "time"):
                try:
                    cell.value = float(cell.value)
                    cell.number_format = "0.00"
                except (TypeError, ValueError):
                    pass

            # Destaque de limite (temp/hum)
            if limits:
                if key in ("avg_temp", "min_temp", "max_temp"):
                    highlight_cell(cell, cell.value, limits.get("temp_min"), limits.get("temp_max"))
                if key in ("avg_hum", "min_hum", "max_hum"):
                    highlight_cell(cell, cell.value, limits.get("hum_min"), limits.get("hum_max"))

    ws.freeze_panes = "A2"
    # Auto ajuste largura
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

def export_one_to_excel(mac, rows, db_manager):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Sensor {mac}"
    headers = [
        ("date", "Data"),
        ("time", "Hora"),
        ("last_temp", "Última Temp (°C)"),
        ("last_hum", "Última Umid (%)"),
        ("last_timestamp", "Hora Última Leitura"),
        ("avg_temp", "Temperatura Média (°C)"),
        ("avg_hum", "Umidade Média (%)"),
        ("min_temp", "Temp. Mín (°C)"),
        ("max_temp", "Temp. Máx (°C)"),
        ("min_hum", "Umid. Mín (%)"),
        ("max_hum", "Umid. Máx (%)"),

    ]
    if not rows:
        ws.append(["Nenhum dado encontrado para o sensor/intervalo selecionado."])
    else:
        limits = get_limits_for_sensor(mac, db_manager)
        write_sensor_sheet(ws, rows, headers, limits)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def export_all_to_excel(json_data, db_manager):
    wb = Workbook()
    wb.remove(wb.active)
    headers = [
        ("date", "Data"),
        ("time", "Hora"),
        ("last_temp", "Última Temp (°C)"),
        ("last_hum", "Última Umid (%)"),
        ("last_timestamp", "Hora Última Leitura"),
        ("avg_temp", "Temperatura Média (°C)"),
        ("avg_hum", "Umidade Média (%)"),
        ("min_temp", "Temp. Mín (°C)"),
        ("max_temp", "Temp. Máx (°C)"),
        ("min_hum", "Umid. Mín (%)"),
        ("max_hum", "Umid. Máx (%)"),
    ]

    for mac, rows in json_data.items():
        ws = wb.create_sheet(title=mac[:31])
        limits = get_limits_for_sensor(mac, db_manager)
        write_sensor_sheet(ws, rows, headers, limits)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

logger = logging.getLogger(__name__)
api_bp = Blueprint('api', __name__, url_prefix='/api/sensors')

def get_payload():
    data = request.get_json(silent=True)
    return data if data is not None else request.form.to_dict()

@api_bp.route('/', methods=['GET'])
def list_sensors():
    """GET /api/sensors/ — lista todos os sensores."""
    sensors = sensor_service.get_all_sensors()
    return jsonify([s.to_dict() for s in sensors]), 200

@api_bp.route('/<mac>', methods=['GET'])
def get_sensor(mac):
    """GET /api/sensors/<mac> — busca metadados de um sensor."""
    sensor = sensor_service.get_sensor(mac)
    if not sensor:
        abort(404)
    return jsonify(sensor.to_dict()), 200

@api_bp.route('/<mac>', methods=['PUT'])
def update_sensor(mac):
    """PUT /api/sensors/<mac> — renomeia um sensor."""
    data = get_payload()
    name = data.get('name')
    if not name:
        return jsonify({"error": "Missing 'name'"}), 400
    sensor_service.rename_sensor(mac, name)
    return jsonify({"status": "ok"}), 200

@api_bp.route('/<mac>/alarms', methods=['GET', 'PUT'])
def sensor_alarms(mac):
    """
    GET  /api/sensors/<mac>/alarms — retorna política de alarmes.
    PUT  /api/sensors/<mac>/alarms — atualiza política de alarmes.
    """
    if request.method == 'GET':
        policy = sensor_service.get_alert_policy(mac)
        return jsonify(policy), 200

    data = get_payload()
    required = ['temp_min','temp_max','humidity_min','humidity_max']
    if not all(k in data for k in required):
        return jsonify({"error": f"Missing one of {required}"}), 400

    result = sensor_service.update_sensor_alert_policy(
        mac,
        temp_min=data['temp_min'],
        temp_max=data['temp_max'],
        humidity_min=data['humidity_min'],
        humidity_max=data['humidity_max']
    )
    return jsonify(result), 200

@api_bp.route('/<mac>/schedules', methods=['GET','PUT'])
def sensor_schedules(mac):
    """
    GET  /api/sensors/<mac>/schedules — retorna política de agendamento.
    PUT  /api/sensors/<mac>/schedules — atualiza política de agendamento.
    """
    if request.method == 'GET':
        sched = sensor_service.get_schedule_policy(mac)
        return jsonify(sched), 200

    data = get_payload()
    delta = data.get('delta_time') or data.get('interval_hours')
    if not delta:
        return jsonify({"error": "Missing 'delta_time'"}), 400

    result = sensor_service.update_sensor_schedule_policy(mac, delta)
    return jsonify(result), 200

@api_bp.route('/<mac>/export', methods=['GET'])
def export_sensor(mac):
    """
    GET /api/sensors/<mac>/export?from=YYYY-MM-DD&to=YYYY-MM-DD&interval=H
    — exporta leituras formatadas (JSON ou CSV).
    """
    fr       = request.args.get('from')
    to       = request.args.get('to')
    interval = request.args.get('interval')
    if not all([fr, to, interval]):
        return jsonify({"error": "Missing parameters"}), 400

    data = sensor_service.export_sensor_data(mac, fr, to, interval)
    return jsonify(data), 200

@api_bp.route('/<mac>/report', methods=['GET'])
def sensor_report(mac):
    """
    GET /api/sensors/<mac>/report?start=...&end=...
    — gera e retorna relatório JSON.
    """
    start = request.args.get('start')
    end   = request.args.get('end')
    if not start or not end:
        return jsonify({"error": "Missing 'start' or 'end'"}), 400

    report = sensor_service.get_sensor_report(mac, start, end)
    return jsonify(report), 200

@api_bp.route('/export_all', methods=['GET'])
def export_all_sensors():
    """
    GET /api/sensors/export_all?from=...&to=...&interval=...
    — retorna agregados de todos os sensores no período.
    """
    fr = request.args.get('from')
    to = request.args.get('to')
    interval = request.args.get('interval')
    if not all([fr, to, interval]):
        return jsonify({"error": "Missing parameters"}), 400

    data = sensor_service.export_all_sensors_data(fr, to, interval)
    return jsonify(data), 200


@api_bp.route('/export_all_excel', methods=['GET'])
def export_all_excel():
    fr = request.args.get('from')
    to = request.args.get('to')
    interval = request.args.get('interval')
    if not all([fr, to, interval]):
        return jsonify({"error": "Missing parameters"}), 400

    data = sensor_service.export_all_sensors_data(fr, to, interval)
    db_manager = sensor_service.db_manager  # <-- PEGUE O DB_MANAGER DO SERVICE
    buf = export_all_to_excel(data, db_manager)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'sensores_{fr}_a_{to}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@api_bp.route('/<mac>/export_excel', methods=['GET'])
def export_sensor_excel(mac):
    fr = request.args.get('from')
    to = request.args.get('to')
    interval = request.args.get('interval')
    if not all([fr, to, interval]):
        return jsonify({"error": "Missing parameters"}), 400

    rows = sensor_service.export_sensor_data(mac, fr, to, interval)
    db_manager = sensor_service.db_manager  # <-- PEGUE O DB_MANAGER DO SERVICE
    buf = export_one_to_excel(mac, rows, db_manager)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'{mac}_{fr}_a_{to}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
